# Standard Python library imports
import glob
import os

# Third party libraries imports
import openpyxl
from openpyxl.drawing.image import Image

# Local module imports
from .my_logger import logger
from .utils import clean_text


class ExcelCreator:
    """
    A class to create Excel files from news data.

    Methods
    -------
    clear_excel_files()
        Clears all existing Excel files in the directory.
    create_excel(news_data)
        Creates an Excel file with the given news data.
    """

    def __init__(self):
        """
        Initializes the ExcelCreator, sets the directory for Excel files, and clears existing files.
        """
        # Here we provide a directory to the code, this directory will or may have some previous documents
        # Those documents are not of our interest right now, so we delete them
        # The function "clear_excel_files()" stablish how we delete the files
        self.excel_files_dir = 'output/'
        os.makedirs(self.excel_files_dir, exist_ok=True)
        self.clear_excel_files()

    def clear_excel_files(self):
        """
        Clears all existing Excel files in the directory.

        Returns
        -------
        None
        """
        logger.info("Starting 'clear_excel_files' function")
        # Here is how we delete all the files
        files = glob.glob(f"{self.excel_files_dir}*.xlsx")
        for f in files:
            os.remove(f)
        logger.info(f"All previous Excel Files from {self.excel_files_dir} directory cleared")

    def create_excel(self, news_data, search_phrase, news_category, num_months):
        """
        Creates a single Excel file containing all the news data.

        Parameters
        ----------
        news_data : list of dict
            List of dictionaries containing the news data.

        Returns
        -------
        None
        """
        logger.info("Starting 'create_excel' function")

        # Check if there are any news items of interest
        if not any(item["bool"] for item in news_data):
            logger.warning("No news items of interest. No Excel file created.")
            return

        # Creating a new Excel Workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Stablishing the Headers
        headers = ["Title", "Date", "Description", "Picture Filename",
                    "Count Search Phrases", "Contains Money"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Initialize row number
        row_num = 2

        # Iterar sobre los datos de las noticias y escribir en el archivo Excel
        for item in news_data:
            # Verify if current new is of interest for me
            if item["bool"]:
                # Write the data in the corresponding cells
                ws.cell(row=row_num, column=1, value=item["title"])
                ws.cell(row=row_num, column=2, value=item["date"])
                ws.cell(row=row_num, column=3, value=item["description"])
                ws.cell(row=row_num, column=4, value=item["image_path"])
                ws.cell(row=row_num, column=5, value=item["phrase_matches"])
                ws.cell(row=row_num, column=6, value=item["contain_money"])

                # Increase the row number for the next news
                row_num += 1


        # Cleaning the search_phrase and the news_category so we can name the file properly
        search_phrase = clean_text(search_phrase)
        news_category = clean_text(news_category)

        # Saving the Excel File
        excel_filename = f"{self.excel_files_dir}{search_phrase}_{news_category}_{num_months}.xlsx"
        wb.save(excel_filename)

        logger.info(f"Single Excel file named {excel_filename} created with all news data")